import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar


window = tk.Tk()
window.title('App de Poupança Pessoal')
window.geometry('700x500')
window.configure(bg='#252525')

style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel",
                background="#252525",
                foreground="#FFFFFF",
                font=("Arial", 12))
style.configure("TEntry",
                fieldbackground="#FFFFFF",
                font=("Arial", 12))
style.configure("TButton",
                background="#4CAF50",
                foreground="#FFFFFF",
                font=("Arial", 12))

valores = []

label_instrucao = ttk.Label(window, text="Insira o valor diário:")
label_status = ttk.Label(window, text="", foreground="red")
label_total = ttk.Label(window, text="", font=("Arial", 14, "bold"))
entry_valor = ttk.Entry(window)
button_salvar = ttk.Button(window, text="Salvar", command=salvar_valor)

# Posicionamento dos elementos
label_instrucao.pack(pady=10)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack()
label_total.pack(pady=10)

workbook = load_workbook('poupanca.xlsx')
sheet = workbook.active

def salvar_valor():
    valor_dia = float(entry_valor.get())
    valores.append(valor_dia)
    total_valores = sum(valores)
    entry_valor.delete(0, tk.END)
    label_status.config(text="Valor salvo com sucesso!", foreground="green")
    label_total.config(text=f"Total economizado: R${total_valores:.2f}")

    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1, value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha, column=2, value=valor_dia)

    canvas.get_tk_widget().destroy()
    plotar_grafico()


def plotar_grafico():
    global canvas

    datas = [cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, "%d-%m-%y").date() for cell in sheet['A'][1:]]
    valores = [cell.value for cell in sheet['B'][1:]]

    dados_mensais = {}
    for data, valor in zip(datas, valores):
        mes_ano = data.strftime("%m-%Y")
        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]

    fig = plt.Figure(figsize=(12, 6), dpi=80)
    ax_barras = fig.add_subplot(121)
    ax_pie = fig.add_subplot(122)

    barras = ax_barras.bar(range(len(dados_mensais)), [sum(valores) for valores in dados_mensais.values()])

    for i, barra in enumerate(barras):
        altura = barra.get_height()
        ax_barras.text(barra.get_x() + barra.get_width() / 2, altura, f'R${altura:.2f}', ha='center', va='bottom')

    nomes_meses = []
    for mes_ano in dados_mensais.keys():
        mes, ano = mes_ano.split('-')
        nome_mes = calendar.month_name[int(mes)]
        nomes_meses.append(f'{nome_mes}-{ano}')

    ax_barras.set_xticks(range(len(dados_mensais)))
    ax_barras.set_xticklabels(nomes_meses, ha='right')

    ax_barras.spines['top'].set_visible(False)
    ax_barras.spines['right'].set_visible(False)
    ax_barras.spines['bottom'].set_visible(False)
    ax_barras.spines['left'].set_visible(False)

    ax_barras.set_title('Economia por Mês')
    ax_barras.title.set_position([.5, 8.05])
    ax_barras.set_xlabel('Mês')
    ax_barras.set_ylabel('Valor Economizado')

    data_inicial = min(datas)
    data_final = max(datas)
    diferenca = (data_final - data_inicial).days
    semanas = diferenca // 7

    labels = [f'{i+1}ª Semana' for i in range(semanas)]
    valores_semana = []
    for i in range(semanas):
        data_inicio = data_inicial + timedelta(weeks=i)
        data_fim = data_inicio + timedelta(weeks=1)
        valores_semana.append(sum(valor for data, valor in zip(datas, valores) if data_inicio <= data < data_fim))

    pie = ax_pie.pie(valores_semana, labels=labels, autopct='%1.1f%%', startangle=90)
    ax_pie.set_title('Economia por Semana')

    canvas = FigureCanvasTkAgg(fig, master=window)
    canvas.get_tk_widget().grid(row=5, column=0, columnspan=2, padx=10, pady=10)

    fig.tight_layout()


plotar_grafico()
window.mainloop()
